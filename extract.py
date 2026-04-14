import PyPDF2
from openai import OpenAI
import numpy as np
import pandas as pd 
import faiss
import streamlit as st
import re
import json
import os

client=OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

def ex_text(f):
    text=""
    df=PyPDF2.PdfReader(f)
    for page in df.pages:
        if page:
            text+=page.extract_text()

    return text


def ai_extract(text):
    prompt=f"""
    You are an expert invoice extractor
    From the data given, extract the values given below
    Data:
    {text}
    Values to be extracted from the given data:
    - Transporter Name
    - LR No
    - Truck No

    - Consignee Name
    - Place(city from consignee address)

    - Supplier Name

    - Line items:
        --Supplier Invoice No
        --Date
        --Total amount

    Strict Rules:
    Do not assume anything
    Extract only what is explicitly present
    Use exact numeric values(no rounding unless present)
    If any field is missing, return null
    Ensure proper spacing between words in names
    
    
    For "Rest of the text" :
    Remove from the data given:
    Remove everything before the texts "DENIER CUTLENGTH MERGE GRADE"
    Similarly remove everything after Cash Discount 
    Return the rest of the text which is in between these two

    

    Output:(Return in this format)
    1.Transporter Name=""
    2.LR No=""
    3.Truck No=""
    4.Name of the Consignee=""
    5.Name of the Supplier=""
    6.Place=""
    7.Supplier Inv No=""
    8.Date=""
    9.Total amount=(dont give with commas)
    10.Rest of the text=""

    """

    ai_res=client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role":"user","content":prompt}]
    )
    ai_out=ai_res.choices[0].message.content

    return ai_out


def reg_ex(text):
    pattern=r"GRADE\s+\d+\s+\d+\s+\d+\s+\d+\s+(\d+\.\d+)\s+(\d+)\s+([A-Z0-9]+)\s+([A-Z]+)\s+(\d+)\s+([\d,]+\.\d+)\s*KG\s*([\d.]+)\s*([\d,]+\.\d+)"
    match=re.findall(pattern,text)
    v=[]
    if match:
        for m in match:
            Denier=float(m[0])
            Cut_length=int(m[1])
            Merge=m[2]
            Grade=m[3]
            No_bales=int(m[4])
            total_kg=float(m[5].replace(",",""))
            basic_rate=float(m[6])
            invoice_amt=float(m[7].replace(",",""))
            a=f"Denier is {Denier},cut length is {Cut_length}, merge is {Merge},grade is {Grade},bales is {No_bales},kgs is {total_kg}, basic rate is {basic_rate} and the invoice amount is {invoice_amt}"
            v.append(a)
        return v
    else:
        return None
    
def change_format(text,v):
    prompt=f"""
    Convert the data given into a structured format given below
    Data:
    {text}
    {v}
    Extra rule:
    If there are multiple values for the date from {v} then display it separately as per the format given below

    Output:
    RETURN as JSON and do not add '''json 
    {{
        "transporter_name":""\n,
        "lr_no":,
        "truck_no":"",

        "supplier_name="",
        "consignee_name="", 
        "place":"",

        "supplier_invoice_no":,
        "invoice_date":"dd-mm-yyyy",
        "total_amount": (dont give commas and give as type float),

        "line_items":[
            "denier": ,
            "cut_length": ,
            "merge":"",
            "grade":"",
            "no_of_bales": ,
            "total_kgs"= ,
            "basic_rate": ,
            "invoice_amount": 
        ]
    }}

    If there are multiple values then mention it as line item 1,2 etc; in the same json format

    """
    response=client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role":"user","content":prompt}]
    )
    output=response.choices[0].message.content
    return output


def fill_in(data):
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    from datetime import datetime
    today_date=datetime.now().strftime("%d-%m-%Y")
    wb=load_workbook("DISPATCH.xlsx")
    count=21
    ws=wb.active
    ws["L14"]=today_date
    ws["L14"].font=Font(bold="True")
    ws["L14"].alignment=Alignment(horizontal="left")
    ws["Q14"]=data["transporter_name"]
    ws["Q14"].font=Font(bold="True")
    ws["Q14"].alignment=Alignment(horizontal="left")
    ws["Q15"]=data["lr_no"]
    ws["Q15"].font=Font(bold="True")
    ws["Q15"].alignment=Alignment(horizontal="left")
    ws["Q16"]=data["truck_no"]
    ws["Q16"].font=Font(bold="True")
    ws["Q16"].alignment=Alignment(horizontal="left")
    ws["Q13"]=data["supplier_name"]
    ws["Q13"].font=Font(bold="True")
    ws["Q13"].alignment=Alignment(horizontal="left")
    ws["L15"]=data["consignee_name"]
    ws["L15"].font=Font(bold="True")
    ws["L15"].alignment=Alignment(horizontal="left")
    ws["L16"]=data["place"]
    ws["L16"].font=Font(bold="True")
    ws["L16"].alignment=Alignment(horizontal="left")
    ws["K21"]=data["supplier_invoice_no"]
    ws["K21"].alignment=Alignment(horizontal="center")
    ws["L21"]=data["invoice_date"]
    ws["L21"].alignment=Alignment(horizontal="center")
    ws["T27"]=(data["total_amount"]+(data["total_amount"]*0.05))
    ws["T27"].font=Font(bold="True")
    ws["T25"]=data["total_amount"]*0.025
    ws["T25"].font=Font(bold="True")
    ws["T26"]=data["total_amount"]*0.025
    ws["T26"].font=Font(bold="True")
    ws["T28"]=((data["total_amount"]+(data["total_amount"]*0.05))*0.001)
    ws["T28"].font=Font(bold="True")
    ws["T30"]=((data["total_amount"]+(data["total_amount"]*0.05))-((data["total_amount"]+(data["total_amount"]*0.05))*0.01))
    ws["T30"].font=Font(bold="True")


    for item in data["line_items"]:

        ws[f"M{count}"]=item["denier"]
        ws[f"N{count}"]=item["cut_length"]
        ws[f"O{count}"]=item["merge"]
        ws[f"P{count}"]=item["grade"]
        ws[f"Q{count}"]=item["no_of_bales"]
        ws[f"R{count}"]=item["total_kgs"]
        ws[f"S{count}"]=item["basic_rate"]
        ws[f"T{count}"]=item["invoice_amount"]
        ws[f"M{count}"].alignment=Alignment(horizontal="center")
        ws[f"N{count}"].alignment=Alignment(horizontal="center")
        ws[f"O{count}"].alignment=Alignment(horizontal="center")
        ws[f"P{count}"].alignment=Alignment(horizontal="center")
        ws[f"Q{count}"].alignment=Alignment(horizontal="center")
        ws[f"R{count}"].alignment=Alignment(horizontal="center")
        ws[f"S{count}"].alignment=Alignment(horizontal="center")
        ws[f"T{count}"].alignment=Alignment(horizontal="center")

        
        
        count+=1
    wb.save("yyy3.xlsx")


        
def st_st():
    st.title("Invoice to Excel Convertor")
    file=st.file_uploader("Upload the File")
    if file:
        st.success("File uploaded successfully - Wait for a few seconds")
        try:
            k=ex_text(file)
            p=ai_extract(k)
            v=reg_ex(p)
            rt=change_format(p,v)
            data=json.loads(rt)
            fill_in(data)
        except:
            st.error("Process Failed")
            return
        with open("yyy3.xlsx","rb") as f:
            st.download_button(
                label="Download",
                data=f,
                file_name="invoice.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )



st_st()